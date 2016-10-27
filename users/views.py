from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse

from .models import Users, Events, Attendances
from .forms import NewUserForm

from datetime import datetime

import pytz
import json
import openpyxl
import tkinter
from tkinter import filedialog

# Create your views here.
def index(request):
    user_list = Users.objects.order_by('id')
    context = {
        'user_list': user_list,
    }
    return render(request, 'users/index.html', context)


# show each user profile
def show(request, user_id):
    user = get_object_or_404(Users, pk=user_id)
    context = {'user': user}

    return render(request, 'users/show.html', context)


# add new user
def new(request):
    context = {}

    if request.method == 'POST':
        form = NewUserForm(request.POST)

        if form.is_valid():
            newUser = Users(name=form.cleaned_data['name'],
                            email=form.cleaned_data['email'],
                            contact=form.cleaned_data['contact'],
                            matric_number=form.cleaned_data['matric_number']
                            )
            newUser.save();
            context['success_message'] = 'New user added'

    context['form'] = NewUserForm()

    return render(request, 'users/new.html', context)

# update user
def update(request):
    if request.method == 'POST':
        dataArray = json.loads(request.POST['dataArray'])
        user_id = request.POST['user_id']

        Users.objects.filter(id=user_id).update(
                name = dataArray[0],
                email = dataArray[1],
                contact = dataArray[2],
                matric_number = dataArray[3],
                updated_at = datetime.now(pytz.timezone('Asia/Singapore'))
            )

        return HttpResponse(
                json.dumps({"isUpdated": True}),
                content_type="application/json"
            )
    else:
        return HttpResponse(
            json.dumps({"nothing to see": "this isn't happening"}),
            content_type="application/json"
        )


# delete user
def delete(request):
    if request.method == 'POST':
        user = get_object_or_404(Users, pk=request.POST['user_id'])
        output = user.delete()
        
        if  output[0]> 0:
            return HttpResponse(
                json.dumps({"isDeleted": True}),
                content_type="application/json"
            )
        else:
            return HttpResponse(
                json.dumps({"isDeleted": False}),
                content_type="application/json"
            )
    else:
        return HttpResponse(
            json.dumps({"nothing to see": "this isn't happening"}),
            content_type="application/json"
        )


# download users
def download(request):
    if request.method == 'POST':

        wb = openpyxl.Workbook()
        ws = wb.active

        # get all members
        memberObjects = Users.objects.all().values()

        colToDelete = []

        # fill up headings
        for count, key in enumerate(memberObjects[0]):
            if (type(memberObjects[0][key]).__name__ != 'datetime'):
                ws.cell(row=1, column=count+2, value=key.upper())
            else:
                colToDelete.append(count)

        # fill up datas
        for count1, memberObject in enumerate(memberObjects):
            for count2, key in enumerate(memberObject):
                if (type(memberObject[key]).__name__ != 'datetime'):
                    ws.cell(row=count1+2, column=count2+2, value=memberObject[key])

                    # ws.write(count1+1, count2, memberObject[key], style1)

        # fill up number colum
        for count in range(len(memberObjects)):
            ws.cell(row=count+2, column=1, value=count+1)

        # delete empty columns for created_at and updated_at
        for count, col in enumerate(colToDelete):
            delete_column(ws, col+2-count)

        # get file path and save
        file_path = get_file_path()

        if not file_path:
            return HttpResponse(
                json.dumps({"downloaded": False}),
                content_type="application/json"
            )
        else:
            wb.save(file_path)
            return HttpResponse(
                    json.dumps({"downloaded": True}),
                    content_type="application/json"
                )
    else:
        return render(request, 'users/download.html')


def attendance(request):
    
    user_list = Users.objects.order_by('id')
    context = {
        'user_list': user_list,
    }
    return render(request, 'users/attendance.html', context)
       


def delete_column(ws, delete_column):
    if isinstance(delete_column, str):
        delete_column = openpyxl.cell.column_index_from_string(delete_column)
    assert delete_column >= 1, "Column numbers must be 1 or greater"

    for column in range(delete_column, ws.max_column + 1):
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=column).value = \
                    ws.cell(row=row, column=column+1).value


def get_file_path():
    root = tkinter.Tk()
    root.attributes('-topmost', True)
    root.withdraw()

    options = {}
    options['defaultextension'] = '.xlsx'
    options['initialfile'] = 'members.xlsx'

    file_path = tkinter.filedialog.asksaveasfilename(**options)

    root.destroy()

    return file_path






    